from __future__ import annotations
import os
import sys
import csv
from datetime import datetime
from typing import List, Dict

from dotenv import load_dotenv
from pymongo import MongoClient
from pymongo.errors import ServerSelectionTimeoutError
import traceback
import certifi

DEFAULT_FILE = os.path.join(os.path.dirname(__file__), 'books.xlsx')


def _try_read_excel(path: str) -> List[Dict]:
    """Try reading Excel via openpyxl. Returns rows as dicts.
    Expects headers: bookid, title, author, price (case-insensitive).
    Raises Exception on failure so caller can fallback.
    """
    try:
        from openpyxl import load_workbook  # lazy import
    except Exception as e:
        raise RuntimeError("openpyxl not installed or failed to import") from e

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]
    out = []
    for r in rows[1:]:
        rec = {}
        for i, v in enumerate(r):
            key = headers[i] if i < len(headers) else f'col{i}'
            rec[key] = v
        out.append(rec)
    return out


essential_keys = {'bookid', 'title', 'author', 'price'}


def _try_read_csv_text(path: str) -> List[Dict]:
    """Fallback: parse the file as CSV text regardless of extension."""
    with open(path, 'r', encoding='utf-8') as f:
        rdr = csv.DictReader(f)
        return list(rdr)


def _norm_row(row: dict) -> dict | None:
    """Normalize a raw row into a book document or return None if invalid."""
    # Flexible key access (case-insensitive)
    lower = { (k or '').strip().lower(): v for k, v in row.items() }
    if not essential_keys.issubset(lower.keys()):
        # Try alternative header names
        # Accept 'book_id' as alias
        if 'book_id' in lower and 'bookid' not in lower:
            lower['bookid'] = lower['book_id']
        if not essential_keys.issubset(lower.keys()):
            return None

    # book_id formatting
    try:
        bid_int = int(str(lower.get('bookid') or '').strip())
        book_id = f"BK-{bid_int:04d}"
    except Exception:
        # If not numeric, keep as-is string
        raw = str(lower.get('bookid') or '').strip()
        if not raw:
            return None
        book_id = raw

    title = str(lower.get('title') or '').strip()
    author = str(lower.get('author') or '').strip()

    # price normalization
    price_raw = lower.get('price')
    try:
        price = float(str(price_raw).strip()) if price_raw is not None and str(price_raw).strip() != '' else None
    except Exception:
        price = None

    if not title:
        return None

    doc = {
        'book_id': book_id,
        'title': title,
        'author': author or 'Unknown',
        'price': price,
        # status defaults to available if new insert; for updates we'll keep existing if present
    }
    return doc


def _connect_db() -> MongoClient:
    """Create a Mongo client quickly with short timeouts. Reads MONGO_URI from .env."""
    load_dotenv()
    mongo_uri = os.environ.get('MONGO_URI') or ''
    if not mongo_uri:
        raise RuntimeError('MONGO_URI not set in environment/.env')
    # Ensure database name 'hostel' exists in URI similar to app.py logic
    if 'hostel' not in mongo_uri:
        if mongo_uri.endswith('?'):
            mongo_uri = mongo_uri + 'retryWrites=true&w=majority&appName=Cluster0'
        mongo_uri = mongo_uri.replace('?', '/hostel?', 1) if '?' in mongo_uri else mongo_uri + '/hostel'

    print("[import] Connecting to MongoDB ...")
    client = MongoClient(
        mongo_uri,
        tlsCAFile=certifi.where(),
        serverSelectionTimeoutMS=3000,
        connectTimeoutMS=3000,
        socketTimeoutMS=3000,
    )
    # Fast ping to validate connectivity
    try:
        client.admin.command('ping')
        print("[import] MongoDB connection OK")
    except ServerSelectionTimeoutError as e:
        raise RuntimeError(f"Failed to connect to MongoDB within timeout. Check internet/firewall and MONGO_URI. Details: {e}")
    except Exception as e:
        raise RuntimeError(f"Failed to connect to MongoDB. Details: {e}")
    return client


def import_books(file_path: str = DEFAULT_FILE) -> dict:
    print(f"[import] Starting import from: {file_path}")
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"File not found: {file_path}")

    # Load rows
    rows = []
    # Try Excel first if extension suggests it
    tried_excel = False
    if os.path.splitext(file_path)[1].lower() in ('.xlsx', '.xlsm', '.xltx', '.xltm'):
        tried_excel = True
        try:
            print("[import] Attempting to read as Excel (openpyxl)...")
            rows = _try_read_excel(file_path)
        except Exception:
            print("[import] Excel read failed, will try CSV fallback")
            rows = []
    # If excel loading failed or ext not excel, try CSV text
    if not rows:
        try:
            print("[import] Reading as CSV text...")
            rows = _try_read_csv_text(file_path)
        except Exception:
            if not tried_excel:
                # Last resort: raise
                raise
            # If both failed, raise a helpful message
            raise RuntimeError("Failed to parse file as Excel or CSV. Ensure headers: bookid,title,author,price")

    # Normalize and upsert
    inserted = 0
    updated = 0
    skipped = 0

    client = _connect_db()
    # Get default DB from URI (should be 'hostel' after the normalization above)
    db = client.get_default_database()
    if db is None:
        db = client['hostel']
    print(f"[import] Parsed {len(rows)} rows. Beginning upsert into 'books' collection...")

    for raw in rows:
        doc = _norm_row(raw)
        if not doc:
            skipped += 1
            continue
        # Try update, else insert
        now = datetime.utcnow()
        existing = db.books.find_one({'book_id': doc['book_id']}, {'_id': 1, 'status': 1})
        if existing:
            # Preserve existing status if set, update other fields
            status = existing.get('status', 'available')
            upd = {k: v for k, v in doc.items()}
            upd['status'] = status
            upd['updated_at'] = now
            res = db.books.update_one({'_id': existing['_id']}, {'$set': upd})
            updated += res.modified_count
        else:
            doc['status'] = 'available'
            doc['created_at'] = now
            db.books.insert_one(doc)
            inserted += 1
    summary = {'inserted': inserted, 'updated': updated, 'skipped': skipped}
    print(f"[import] Done. Summary: {summary}")
    return summary


if __name__ == '__main__':
    try:
        path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_FILE
        result = import_books(path)
        print(f"Import complete. Inserted: {result['inserted']}, Updated: {result['updated']}, Skipped: {result['skipped']}")
    except Exception as e:
        print("[import] ERROR:")
        print(e)
        traceback.print_exc()
        sys.exit(1)
